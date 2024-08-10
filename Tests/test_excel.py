import unittest
import os
import pandas as pd
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from Functions.models import Clients, Client, Sales, Sale, Workers, Worker, Expenses, Expense
from Functions.excel import SaveAllInExcel, LoadAllFromTable, SaveOneInExcel

class TestSaveAllInExcel(unittest.TestCase):
    def setUp(self):
        self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.excel_file_name = self.test_excel_file.name

        self.test_db_file = NamedTemporaryFile(delete=False)
        self.db_file_name = self.test_db_file.name
        self.test_db_file.close()

        Clients.create_table_client(self.db_file_name)
        clients_data = [
            Client('M6777-0', 'Mario', 'Castañeda', 'M', 40, 'mariocasta@gmail.com'),
            Client('70000-1', 'María', 'Liendra', 'F', 32, 'liendma@gmail.com')
        ]
        Clients.add_many_clients(clients_data, self.db_file_name)

        Sales.create_table_sale(self.db_file_name)
        sales_data = [
            Sale('10000000J', None, '2023-12-10 12:30:00', 50, 0, 1, 'X20 Hot Dogs'),
            Sale('39000000K', None, '2023-12-10 21:45:00', 15, 0, 1, 'X1 Bowl of Nuggets')
        ]
        Sales.add_many_sales(sales_data, self.db_file_name)

        Workers.create_table_worker(self.db_file_name)
        workers_data = [
            Worker('00000-1', 'Viviana', 'Marcic', 'Executive', 6000, 'v.marcic@gmail.com'),
            Worker('M9000-1', 'Eugene', 'Parker', 'Waiter', 150, 'eug@gmail.com')
        ]
        Workers.add_many_workers(workers_data, self.db_file_name)

        Expenses.create_table_expense(self.db_file_name)
        expenses_data = [
            Expense('05050505L', None, '2023-12-20 09:45:00', 1750, 0, 'News salarys'),
            Expense('45009000K', None, '2022-01-23 21:45:00', 50, 1, 'Buy knifes')
        ]
        Expenses.add_many_expenses(expenses_data, self.db_file_name)

    def tearDown(self):
        self.test_excel_file.close()
        try:
            os.remove(self.db_file_name)
            os.remove(self.excel_file_name)
        except PermissionError:
            pass

    def test_save_all_clients_in_excel(self):
        SaveAllInExcel.save_all_clients_in_excel(self.excel_file_name, self.db_file_name)
        new_clients_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Clients', engine='openpyxl')
        self.assertEqual(len(new_clients_in_excel), 2)
        self.assertEqual(new_clients_in_excel['name'].to_list(), ['Mario', 'María'])

    def test_save_all_sales_in_excel(self):
        SaveAllInExcel.save_all_sales_in_excel(self.excel_file_name, self.db_file_name)
        new_sales_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Sales', engine='openpyxl')
        client_ids = new_sales_in_excel['client_id'].apply(lambda x: None if pd.isna(x) else x).to_list()
        self.assertEqual(len(new_sales_in_excel), 2)
        self.assertEqual(client_ids, [None, None])

    def test_save_all_workers_in_excel(self):
        SaveAllInExcel.save_all_workers_in_excel(self.excel_file_name, self.db_file_name)
        new_workers_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Workers', engine='openpyxl')
        self.assertEqual(len(new_workers_in_excel), 2)
        self.assertEqual(new_workers_in_excel['salary'].to_list(), [6000, 150])

    def test_save_all_expenses_in_excel(self):
        SaveAllInExcel.save_all_expenses_in_excel(self.excel_file_name, self.db_file_name)
        new_expenses_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Expenses', engine='openpyxl')
        self.assertEqual(len(new_expenses_in_excel), 2)
        self.assertEqual(new_expenses_in_excel['paid'].to_list(), [0, 1])

class TestLoadAllFromTable(unittest.TestCase):
    def setUp(self):
        self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.excel_file_name = self.test_excel_file.name
        
        clients_data = {
            'id': ['M6777-0', '70000-1'],
            'name': ['Mario', 'María'],
            'surname': ['Castañeda', 'Liendra'],
            'gender': ['M', 'F'],
            'age': [40, 32],
            'mail': ['mariocasta@gmail.com', 'liendma@gmail.com']
        }
        df_clients = pd.DataFrame(clients_data)
        
        sales_data = {
            'id': ['10000007G', '39000000J'],
            'client_id': [None, None],
            'date': ['2023-12-10 12:30:00', '2023-12-10 21:45:00'],
            'cash': [50, 15],
            'paid': [0, 0],
            'delivered': [1, 1],
            'description': ['X20 Hot Dogs', 'X1 Bowl of Nuggets']
        }
        df_sales = pd.DataFrame(sales_data)

        workers_data = {
            'id': ['00000-1', 'M9000-1'],
            'name': ['Viviana', 'Eugene'],
            'surname': ['Marcic', 'Parker'],
            'position': ['Executive', 'Waiter'],
            'salary': [6000, 150],
            'mail': ['v.marcic@gmail.com', 'eug@gmail.com']
        }
        df_workers = pd.DataFrame(workers_data)

        expenses_data = {
            'id': ['05000000Y', '45000053D'],
            'worker_id': [None, None],
            'date': ['2023-12-20 09:45:00', '2022-01-23 21:45:00'],
            'cash': [1750, 50],
            'paid': [0, 1],
            'description': ['News salarys', 'Buy knifes']
        }
        df_expenses = pd.DataFrame(expenses_data)

        with pd.ExcelWriter(self.excel_file_name, engine='openpyxl') as writer:
            df_clients.to_excel(writer, sheet_name='Clients', index=False)
            df_sales.to_excel(writer, sheet_name='Sales', index=False)
            df_workers.to_excel(writer, sheet_name='Workers', index=False)
            df_expenses.to_excel(writer, sheet_name='Expenses', index=False)

    def tearDown(self):
        self.test_excel_file.close()
        try:
            os.remove(self.excel_file_name)
        except PermissionError:
            pass

    def test_load_clients_from_excel(self):
        clients = LoadAllFromTable.load_clients_from_excel(self.excel_file_name)
        self.assertEqual(len(clients), 2)
        self.assertEqual(clients[0].name, 'Mario')
        self.assertEqual(clients[1].name, 'María')

    def test_load_sales_from_excel(self):
        sales = LoadAllFromTable.load_sales_from_excel(self.excel_file_name)
        self.assertEqual(len(sales), 2)
        self.assertIsNone(sales[0].client_id)
        self.assertIsNone(sales[1].client_id)

    def test_load_workers_from_excel(self):
        workers = LoadAllFromTable.load_workers_from_excel(self.excel_file_name)
        self.assertEqual(len(workers), 2)
        self.assertEqual(workers[0].salary, 6000)
        self.assertEqual(workers[1].salary, 150)

    def test_load_expenses_from_excel(self):
        expenses = LoadAllFromTable.load_expenses_from_excel(self.excel_file_name)
        self.assertEqual(len(expenses), 2)
        self.assertEqual(expenses[0].paid, 0)
        self.assertEqual(expenses[1].paid, 1)

    def test_load_all_from_excel(self):
        clients, sales, workers, expenses = LoadAllFromTable.load_all_from_excel(self.excel_file_name)
        self.assertEqual(len(clients), 2)
        self.assertEqual(clients[0].name, 'Mario')
        self.assertEqual(clients[1].name, 'María')

        self.assertEqual(len(sales), 2)
        self.assertIsNone(sales[0].client_id)
        self.assertIsNone(sales[1].client_id)

        self.assertEqual(len(workers), 2)
        self.assertEqual(workers[0].salary, 6000)
        self.assertEqual(workers[1].salary, 150)

        self.assertEqual(len(expenses), 2)
        self.assertEqual(expenses[0].paid, 0)
        self.assertEqual(expenses[1].paid, 1)

class TestSaveOneInExcel(unittest.TestCase):
    def setUp(self):
        self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.excel_file_name = self.test_excel_file.name

        wb = Workbook()
        wb.create_sheet(title="Clients")
        wb.create_sheet(title="Sales")
        wb.create_sheet(title="Workers")
        wb.create_sheet(title="Expenses")
        wb.remove(wb["Sheet"])  
        wb.save(self.excel_file_name)

        self.test_client = Client('70000-1', 'María', 'Liendra', 'F', 32, 'liendma@gmail.com')
        self.test_sale = Sale('39000000E', None, '2023-12-10 21:45:00', 15, 0, 1, 'X1 Bowl of Nuggets')
        self.test_worker = Worker('00000-1', 'Viviana', 'Marcic', 'Executive', 6000, 'v.marcic@gmail.com')
        self.test_expense = Expense('45000000W', None, '2022-01-23 21:45:00', 50, 1, 'Buy knifes')

    def tearDown(self):
        self.test_excel_file.close()
        try:
            os.remove(self.excel_file_name)
        except PermissionError:
            pass

    def test_save_client_in_excel(self):
        SaveOneInExcel.save_client_in_excel(self.excel_file_name, self.test_client)
        new_client_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Clients', engine='openpyxl')
        self.assertEqual(len(new_client_in_excel), 1)
        self.assertEqual(new_client_in_excel.loc[0, 'name'], 'María')

    def test_save_sale_in_excel(self):
        SaveOneInExcel.save_sale_in_excel(self.excel_file_name, self.test_sale)
        new_sale_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Sales', engine='openpyxl')
        client_id = new_sale_in_excel['client_id'].apply(lambda x: None if pd.isna(x) else x)
        self.assertEqual(len(new_sale_in_excel), 1)
        self.assertIsNone(client_id[0])

    def test_save_worker_in_excel(self):
        SaveOneInExcel.save_worker_in_excel(self.excel_file_name, self.test_worker)
        new_worker_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Workers', engine='openpyxl')
        self.assertEqual(len(new_worker_in_excel), 1)
        self.assertEqual(new_worker_in_excel.loc[0, 'salary'], 6000)

    def test_save_expense_in_excel(self):
        SaveOneInExcel.save_expense_in_excel(self.excel_file_name, self.test_expense)
        new_expense_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Expenses', engine='openpyxl')
        self.assertEqual(len(new_expense_in_excel), 1)
        self.assertEqual(new_expense_in_excel.loc[0, 'paid'], 1)