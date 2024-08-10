from copy import copy
from tempfile import NamedTemporaryFile
import pandas as pd
from datetime import datetime
import os
import unittest
from openpyxl import Workbook
from Functions.models import Client, Clients, Sale, Sales, Worker, Workers, Expense, Expenses, Tasks
from Functions.excel import SaveAllInExcel, SaveOneInExcel, LoadAllFromTable
from Functions.info import InfoClients, InfoSales
from Functions.Auxiliars.valid_id import valid_id

class TestModels:
    class TestModelClient(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Clients.create_table_client(self.db_file_name)
            Clients.add_client('ABCDE-1', 'John', 'Doe', 'M', 30, 'tr@gmail.com', self.db_file_name)
            Clients.add_client('XYZ00-2', 'Alice', 'Smith', 'F', 25, 'y@gmail.com', self.db_file_name)
            Clients.add_client('12345-3', 'Michael', 'Johnson', 'M', 40, 'urs@gmail.com', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_search_client(self):
            true_client = Clients.search_client('ABCDE-1', self.db_file_name)
            false_client = Clients.search_client('00000-0', self.db_file_name)
            self.assertIsNotNone(true_client)
            self.assertIsNone(false_client)

        def test_add_client(self):
            previous_list_client = copy(Clients.load_clients(self.db_file_name))
            new_client = Clients.add_client('PLG00-H', 'Marie', 'Curie', 'F', 40, 'l@gmail.com', self.db_file_name)
            new_list_clients = Clients.load_clients(self.db_file_name)
            new_client_in_db = Clients.search_client('PLG00-H', self.db_file_name)
            diferrence_totals = len(new_list_clients) - len(previous_list_client)
            self.assertEqual(diferrence_totals, 1)
            self.assertEqual(new_client.id, 'PLG00-H')
            self.assertEqual(new_client.name, 'Marie')
            self.assertEqual(new_client.surname, 'Curie')
            self.assertEqual(new_client.gender, 'F')
            self.assertEqual(new_client.age, 40)
            self.assertEqual(new_client.name, new_client_in_db.name)

        def test_modificate_client(self):
            client_to_modificate = copy(Clients.search_client('XYZ00-2', self.db_file_name))
            modificated_client = Clients.modificate_client('XYZ00-2', 'Charlotte', 'Smith', 'F', 25, 'sa@gmail.com', self.db_file_name)
            new_client = Clients.search_client('XYZ00-2', self.db_file_name)
            self.assertEqual(client_to_modificate.name, 'Alice')
            self.assertEqual(modificated_client.name, 'Charlotte')
            self.assertEqual(new_client.name, 'Charlotte')

        def test_remove_client(self):
            previous_list_clients = copy(Clients.load_clients(self.db_file_name))
            removed_client = Clients.remove_client('12345-3', self.db_file_name)
            search_removed_client = Clients.search_client('12345-3', self.db_file_name)
            difference_totals = len(previous_list_clients) - len(Clients.load_clients(self.db_file_name))
            self.assertEqual(removed_client.id, '12345-3')
            self.assertIsNone(search_removed_client)
            self.assertEqual(difference_totals, 1)

        def test_add_many_clients(self):
            previous_list_clients = copy(Clients.load_clients(self.db_file_name))
            list_of_new_clients = [Client('DEFGH-4', 'Emily', 'Brown', 'F', 35, 'cde@gmail.com'),
                                Client('45678-A', 'James', 'Williams', 'M', 28, 'efvg@gmail.com')]
            Clients.add_many_clients(list_of_new_clients, self.db_file_name)
            difference_totals = len(Clients.load_clients(self.db_file_name)) - len(previous_list_clients)
            self.assertEqual(difference_totals, 2)

        def test_remove_all_clients(self):
            Clients.remove_all_clients(self.db_file_name)
            new_register = Clients.load_clients(self.db_file_name)
            empty_list = len(new_register)
            self.assertEqual(empty_list, 0)

    class TestModelSale(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Sales.create_table_sale(self.db_file_name)
            Clients.create_table_client(self.db_file_name)
            Clients.add_client('NJ876-9', 'John', 'Doe', 'M', 30, 'as@gmail.com', self.db_file_name)
            Sales.add_sale('00011123K', 'NJ876-9', '2024-05-10 21:30:00', 150, 1, 1, 'X5 Pasta Bolognese Combo', self.db_file_name)
            Sales.add_sale('70770707K', 'NJ876-9', '2024-01-01 12:00:00', 50, 0, 1, 'X1 Large Neapolitan Pizza & X1 Large Jam Pizza', self.db_file_name)
            Sales.add_sale('06440400G', 'NJ876-9', '2024-02-04 12:30:00', 10, 0, 0, 'X1 Caprese Salad', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_search_sale(self):
            true_sale = Sales.search_sale('00011123K', self.db_file_name)
            false_sale = Sales.search_sale('33333333J', self.db_file_name)
            self.assertIsNotNone(true_sale)
            self.assertIsNone(false_sale)

        def test_add_sale(self):
            previous_list_sale = copy(Sales.load_sales(self.db_file_name))
            new_sale = Sales.add_sale('00000000D', 'NJ876-9', '27/01/2024', 400, 0, 0, 'X3 Large Sicilian Pizza & X1 Tignanello Bottle', self.db_file_name)
            new_sale_in_db = Sales.search_sale('00000000D', self.db_file_name)
            new_list_sale = Sales.load_sales(self.db_file_name)
            difference_totals = len(new_list_sale) - len(previous_list_sale)
            self.assertEqual(difference_totals, 1)
            self.assertEqual(new_sale.id, '00000000D')
            self.assertEqual(new_sale.client_id, 'NJ876-9')
            self.assertEqual(new_sale.date, '27/01/2024')
            self.assertEqual(new_sale.cash, 400)
            self.assertEqual(new_sale.paid, 0)
            self.assertEqual(new_sale.delivered, 0)
            self.assertEqual(new_sale.cash, new_sale_in_db.cash)

        def test_modificate_sale(self):
            sale_to_modificate = copy(Sales.search_sale('70770707K', self.db_file_name))
            modificated_sale = Sales.modificate_sale('70770707K', 25, 1, 0, 'X1 Large Jam Pizza', self.db_file_name)
            new_sale = Sales.search_sale('70770707K', self.db_file_name)
            self.assertEqual(sale_to_modificate.cash, 50)
            self.assertEqual(modificated_sale.cash, 25)
            self.assertEqual(new_sale.cash, 25)

        def test_remove_sale(self):
            previous_list_sales = copy(Sales.load_sales(self.db_file_name))
            removed_sale = Sales.remove_sale('06440400G', self.db_file_name)
            search_removed_sale = Sales.search_sale('06440400G', self.db_file_name)
            difference_totals = len(previous_list_sales) - len(Sales.load_sales(self.db_file_name))
            self.assertEqual(removed_sale.id, '06440400G')
            self.assertIsNone(search_removed_sale)
            self.assertEqual(difference_totals, 1)

        def test_add_many_sales(self):
            previous_list_sales = copy(Sales.load_sales(self.db_file_name))
            list_of_new_sales = [Sale('65889888Q', 'NJ876-9', '12/07/2023', 90, 1, 1, 'X1 Gnocchi Arrabbiata & X1 Ossobuco'),
                                Sale('01363610N', 'NJ876-9', '12/07/2023', 23, 0, 0, 'X1 Castagnaccio & X1 Ristretto')]
            Sales.add_many_sales(list_of_new_sales, self.db_file_name)
            difference_totals = len(Sales.load_sales(self.db_file_name)) - len(previous_list_sales)
            self.assertEqual(difference_totals, 2)

        def test_remove_all_sales(self):
            Sales.remove_all_sales(self.db_file_name)
            new_register = Sales.load_sales(self.db_file_name)
            empty_list = len(new_register)
            self.assertEqual(empty_list, 0)

    class TestModelWoker(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Workers.create_table_worker(self.db_file_name)
            Workers.add_worker('OPL55-5', 'Carlos', 'Quintero', 'Cashier', 550, 'plm@gmail.com', self.db_file_name)
            Workers.add_worker('30111-J', 'Nerea', 'Linares', 'Waitress', 750, 'yt@gmail.com', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_search_worker(self):
            true_worker = Workers.search_worker('OPL55-5', self.db_file_name)
            false_worker = Workers.search_worker('00000-0', self.db_file_name)
            self.assertIsNotNone(true_worker)
            self.assertIsNone(false_worker)

        def test_add_worker(self):
            previous_list_worker = copy(Workers.load_workers(self.db_file_name))
            new_worker = Workers.add_worker('C4004-0', 'Crystal', 'Moreno', 'Manager', 1500, 'oe@gmail.com', self.db_file_name)
            new_list_workers = Workers.load_workers(self.db_file_name)
            new_worker_in_db = Workers.search_worker('C4004-0', self.db_file_name)
            difference_totals = len(new_list_workers) - len(previous_list_worker)
            self.assertEqual(difference_totals, 1)
            self.assertEqual(new_worker.id, 'C4004-0')
            self.assertEqual(new_worker.name, 'Crystal')
            self.assertEqual(new_worker.surname, 'Moreno')
            self.assertEqual(new_worker.position, 'Manager')
            self.assertEqual(new_worker.salary, 1500)
            self.assertEqual(new_worker.name, new_worker_in_db.name)

        def test_modificate_worker(self):
            worker_to_modificate = copy(Workers.search_worker('30111-J', self.db_file_name))
            modificated_client = Workers.modificate_worker('30111-J', 'Maria', 'Linares', 'Waitress', 750, 'res@gmail.com', self.db_file_name)
            new_worker = Workers.search_worker('30111-J', self.db_file_name)
            self.assertEqual(worker_to_modificate.name, 'Nerea')
            self.assertEqual(modificated_client.name, 'Maria')
            self.assertEqual(new_worker.name, 'Maria')

        def test_remove_worker(self):
            previous_list_workers = copy(Workers.load_workers(self.db_file_name))
            removed_worker = Workers.remove_worker('30111-J', self.db_file_name)
            search_removed_worked = Workers.search_worker('30111-J', self.db_file_name)
            difference_totals = len(previous_list_workers) - len(Workers.load_workers(self.db_file_name))
            self.assertEqual(removed_worker.id, '30111-J', self.db_file_name)
            self.assertIsNone(search_removed_worked)
            self.assertEqual(difference_totals, 1)

        def test_add_many_workers(self):
            previous_list_workers = copy(Workers.load_workers(self.db_file_name))
            list_of_new_workers = [Worker('WARRR-0', 'Tyron', 'Valenzuela', 'Chef', 1500, 'pyg@gmail.com'), 
                                   Worker('PAZZZ-0', 'Jenny', 'Rauch', 'Kitchen assistant', 200, 'ytd@gmail.com')]
            Workers.add_many_workers(list_of_new_workers, self.db_file_name)
            difference_totals = len(Workers.load_workers(self.db_file_name)) - len(previous_list_workers)
            self.assertEqual(difference_totals, 2)
            
        def test_remove_all_workers(self):
            Workers.remove_all_workers(self.db_file_name)
            new_register = Workers.load_workers(self.db_file_name)
            empty_list = len(new_register)
            self.assertEqual(empty_list, 0)

    class TestModelExpense(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Expenses.create_table_expense(self.db_file_name)
            Workers.create_table_worker(self.db_file_name)
            Workers.add_worker('IUI00-0', 'Marco', 'Polo', 'Waiter', 21, 'jdsjdjs@gmail.com', self.db_file_name)
            Expenses.add_expense('22290980Q', 'IUI00-0', '2021-10-12 08:21:00', 500, 1, 'Salary of Marco', self.db_file_name)
            Expenses.add_expense('32000000T', None, '2008-02-03 20:21:00', 30, 0, 'Recompose a client', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_search_expense(self):
            true_expense = Expenses.search_expense('22290980Q', self.db_file_name)
            false_expense = Expenses.search_expense('01002037A', self.db_file_name)
            self.assertIsNotNone(true_expense)
            self.assertIsNone(false_expense)

        def test_add_expense(self):
            previous_list_expense = copy(Expenses.load_expenses(self.db_file_name))
            new_expense = Expenses.add_expense('32000213R', None, '2023-12-01 10:30:00', 5000, 1, 'Pay bills', self.db_file_name)
            new_expense_in_db = Expenses.search_expense('32000213R', self.db_file_name)
            new_list_expense = Expenses.load_expenses(self.db_file_name)
            difference_totals = len(new_list_expense) - len(previous_list_expense)
            self.assertEqual(difference_totals, 1)
            self.assertEqual(new_expense.id, '32000213R')
            self.assertEqual(new_expense.worker_id, None)
            self.assertEqual(new_expense.date, '2023-12-01 10:30:00')
            self.assertEqual(new_expense.cash, 5000)
            self.assertEqual(new_expense.paid, 1)
            self.assertEqual(new_expense.cash, new_expense_in_db.cash)

        def test_modificate_expense(self):
            expense_to_modificate = copy(Expenses.search_expense('32000000T', self.db_file_name))
            modificated_expense = Expenses.modificate_expense('32000000T', '2021-10-12 08:21:00', 600, 1, 'Salary of Marco', self.db_file_name)
            new_expense = Expenses.search_expense('32000000T', self.db_file_name)
            self.assertEqual(expense_to_modificate.cash, 30)
            self.assertEqual(modificated_expense.cash, 600)
            self.assertEqual(new_expense.cash, 600)

        def test_remove_expense(self):
            previous_list_expense = copy(Expenses.load_expenses(self.db_file_name))
            removed_expense = Expenses.remove_expense('32000000T', self.db_file_name)
            searc_removed_expense = Expenses.search_expense('32000000T', self.db_file_name)
            difference_totals = len(previous_list_expense) - len(Expenses.load_expenses(self.db_file_name))
            self.assertEqual(removed_expense.id, '32000000T')
            self.assertIsNone(searc_removed_expense)
            self.assertEqual(difference_totals, 1)

        def test_add_many_expenses(self):
            previous_list_expenses = copy(Expenses.load_expenses(self.db_file_name))
            list_of_new_expenses = [Expense('20000000E', None, '2019-09-23 10:05:00', 81, 0, 'Replenish ingredients'),
                                    Expense('50000000G', None, '2019-10-23 10:05:00', 100, 0, 'Buy utensils')]
            Expenses.add_many_expenses(list_of_new_expenses, self.db_file_name)
            difference_totals = len(Expenses.load_expenses(self.db_file_name)) - len(previous_list_expenses)
            self.assertEqual(difference_totals, 2)

        def test_remove_all_expenses(self):
            Expenses.remove_all_expenses(self.db_file_name)
            new_register = Expenses.load_expenses(self.db_file_name)
            empty_list = len(new_register)
            self.assertEqual(empty_list, 0)

    class TestModelTask(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Tasks.create_table_task(self.db_file_name)
            Workers.create_table_worker(self.db_file_name)
            Workers.add_worker('IUI00-0', 'Marco', 'Polo', 'Waiter', 21, 'jdsjdjs@gmail.com', self.db_file_name)
            Tasks.add_task('TTTYGTFR6', 'IUI00-0', 'clean dishes', 'blablabla', 0, self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_search_task(self):
            true_task = Tasks.search_task('TTTYGTFR6', self.db_file_name)
            false_task = Tasks.search_task('NNHFHJHH7', self.db_file_name)

            self.assertEqual(true_task.id, 'TTTYGTFR6')
            self.assertEqual(true_task.worker_id, 'IUI00-0')
            self.assertEqual(true_task.title, 'clean dishes')
            self.assertEqual(true_task.content, 'blablabla')
            self.assertEqual(true_task.done, 0)
            self.assertIsNone(false_task)

        def test_add_task(self):
            previous_list_task = copy(Tasks.load_tasks(self.db_file_name))
            new_task = Tasks.add_task('RRRRRRRR5', None, 'Cook', 'blabla', 1, self.db_file_name)
            new_task_in_db = Tasks.search_task('RRRRRRRR5', self.db_file_name)
            new_list_task = Tasks.load_tasks(self.db_file_name)
            difference_totals = len(new_list_task) - len(previous_list_task)
            self.assertEqual(difference_totals, 1)
            self.assertEqual(new_task.id, 'RRRRRRRR5')
            self.assertEqual(new_task.worker_id, None)
            self.assertEqual(new_task.title, 'Cook')
            self.assertEqual(new_task.content, 'blabla')
            self.assertEqual(new_task.done, 1)
            self.assertEqual(new_task.title, new_task_in_db.title)

        def test_modificate_task(self):
            task_to_modificate = copy(Tasks.search_task('TTTYGTFR6', self.db_file_name))
            modificated_task = Tasks.modificate_task('TTTYGTFR6', None, 'Cook', 'blabla', 1, self.db_file_name)
            new_task = Tasks.search_task('TTTYGTFR6', self.db_file_name)
            self.assertEqual(task_to_modificate.content, 'blablabla')
            self.assertEqual(modificated_task.content, 'blabla')
            self.assertEqual(new_task.content, 'blabla')

        def test_remove_task(self):
            previous_list_task = copy(Tasks.load_tasks(self.db_file_name))
            removed_task = Tasks.remove_task('TTTYGTFR6', self.db_file_name)
            search_removed_task = Tasks.search_task('TTTYGTFR6', self.db_file_name)
            difference_totals = len(previous_list_task) - len(Tasks.load_tasks(self.db_file_name))
            self.assertEqual(removed_task.id, 'TTTYGTFR6')
            self.assertIsNone(search_removed_task)
            self.assertEqual(difference_totals, 1)

        def test_remove_all_tasks(self):
            Tasks.remove_all_task(self.db_file_name)
            new_register = Tasks.load_tasks(self.db_file_name)
            empty_list = len(new_register)
            self.assertEqual(empty_list, 0)

class TestExcel:
    class TestSaveAllInExcel(unittest.TestCase):
        def setUp(self):
            self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
            self.excel_file_name = self.test_excel_file.name

            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()

            Clients.create_table_client(self.db_file_name)
            clients_data = [
                Client('M6777-0', 'Mario', 'Castañeda', 'M', 40, 'mariocasta@gmail.com'),
                Client('70000-1', 'María', 'Liendra', 'F', 32, 'liendma@gmail.com')
            ]
            Clients.add_many_clients(clients_data, self.db_file_name)

            Sales.create_table_sale(self.db_file_name)
            sales_data = [
                Sale('10000000J', None, '2023-12-10 12:30:00', 50, 0, 1, 'X20 Hot Dogs'),
                Sale('39000000K', None, '2023-12-10 21:45:00', 15, 0, 1, 'X1 Bowl of Nuggets')
            ]
            Sales.add_many_sales(sales_data, self.db_file_name)

            Workers.create_table_worker(self.db_file_name)
            workers_data = [
                Worker('00000-1', 'Viviana', 'Marcic', 'Executive', 6000, 'v.marcic@gmail.com'),
                Worker('M9000-1', 'Eugene', 'Parker', 'Waiter', 150, 'eug@gmail.com')
            ]
            Workers.add_many_workers(workers_data, self.db_file_name)

            Expenses.create_table_expense(self.db_file_name)
            expenses_data = [
                Expense('05050505L', None, '2023-12-20 09:45:00', 1750, 0, 'News salarys'),
                Expense('45009000K', None, '2022-01-23 21:45:00', 50, 1, 'Buy knifes')
            ]
            Expenses.add_many_expenses(expenses_data, self.db_file_name)

        def tearDown(self):
            self.test_excel_file.close()
            try:
                os.remove(self.db_file_name)
                os.remove(self.excel_file_name)
            except PermissionError:
                pass

        def test_save_all_clients_in_excel(self):
            SaveAllInExcel.save_all_clients_in_excel(self.excel_file_name, self.db_file_name)
            new_clients_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Clients', engine='openpyxl')
            self.assertEqual(len(new_clients_in_excel), 2)
            self.assertEqual(new_clients_in_excel['name'].to_list(), ['Mario', 'María'])

        def test_save_all_sales_in_excel(self):
            SaveAllInExcel.save_all_sales_in_excel(self.excel_file_name, self.db_file_name)
            new_sales_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Sales', engine='openpyxl')
            client_ids = new_sales_in_excel['client_id'].apply(lambda x: None if pd.isna(x) else x).to_list()
            self.assertEqual(len(new_sales_in_excel), 2)
            self.assertEqual(client_ids, [None, None])

        def test_save_all_workers_in_excel(self):
            SaveAllInExcel.save_all_workers_in_excel(self.excel_file_name, self.db_file_name)
            new_workers_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Workers', engine='openpyxl')
            self.assertEqual(len(new_workers_in_excel), 2)
            self.assertEqual(new_workers_in_excel['salary'].to_list(), [6000, 150])

        def test_save_all_expenses_in_excel(self):
            SaveAllInExcel.save_all_expenses_in_excel(self.excel_file_name, self.db_file_name)
            new_expenses_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Expenses', engine='openpyxl')
            self.assertEqual(len(new_expenses_in_excel), 2)
            self.assertEqual(new_expenses_in_excel['paid'].to_list(), [0, 1])

    class TestLoadAllFromTable(unittest.TestCase):
        def setUp(self):
            self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
            self.excel_file_name = self.test_excel_file.name
            
            clients_data = {
                'id': ['M6777-0', '70000-1'],
                'name': ['Mario', 'María'],
                'surname': ['Castañeda', 'Liendra'],
                'gender': ['M', 'F'],
                'age': [40, 32],
                'mail': ['mariocasta@gmail.com', 'liendma@gmail.com']
            }
            df_clients = pd.DataFrame(clients_data)
            
            sales_data = {
                'id': ['10000007G', '39000000J'],
                'client_id': [None, None],
                'date': ['2023-12-10 12:30:00', '2023-12-10 21:45:00'],
                'cash': [50, 15],
                'paid': [0, 0],
                'delivered': [1, 1],
                'description': ['X20 Hot Dogs', 'X1 Bowl of Nuggets']
            }
            df_sales = pd.DataFrame(sales_data)

            workers_data = {
                'id': ['00000-1', 'M9000-1'],
                'name': ['Viviana', 'Eugene'],
                'surname': ['Marcic', 'Parker'],
                'position': ['Executive', 'Waiter'],
                'salary': [6000, 150],
                'mail': ['v.marcic@gmail.com', 'eug@gmail.com']
            }
            df_workers = pd.DataFrame(workers_data)

            expenses_data = {
                'id': ['05000000Y', '45000053D'],
                'worker_id': [None, None],
                'date': ['2023-12-20 09:45:00', '2022-01-23 21:45:00'],
                'cash': [1750, 50],
                'paid': [0, 1],
                'description': ['News salarys', 'Buy knifes']
            }
            df_expenses = pd.DataFrame(expenses_data)

            with pd.ExcelWriter(self.excel_file_name, engine='openpyxl') as writer:
                df_clients.to_excel(writer, sheet_name='Clients', index=False)
                df_sales.to_excel(writer, sheet_name='Sales', index=False)
                df_workers.to_excel(writer, sheet_name='Workers', index=False)
                df_expenses.to_excel(writer, sheet_name='Expenses', index=False)

        def tearDown(self):
            self.test_excel_file.close()
            try:
                os.remove(self.excel_file_name)
            except PermissionError:
                pass

        def test_load_clients_from_excel(self):
            clients = LoadAllFromTable.load_clients_from_excel(self.excel_file_name)
            self.assertEqual(len(clients), 2)
            self.assertEqual(clients[0].name, 'Mario')
            self.assertEqual(clients[1].name, 'María')

        def test_load_sales_from_excel(self):
            sales = LoadAllFromTable.load_sales_from_excel(self.excel_file_name)
            self.assertEqual(len(sales), 2)
            self.assertIsNone(sales[0].client_id)
            self.assertIsNone(sales[1].client_id)

        def test_load_workers_from_excel(self):
            workers = LoadAllFromTable.load_workers_from_excel(self.excel_file_name)
            self.assertEqual(len(workers), 2)
            self.assertEqual(workers[0].salary, 6000)
            self.assertEqual(workers[1].salary, 150)

        def test_load_expenses_from_excel(self):
            expenses = LoadAllFromTable.load_expenses_from_excel(self.excel_file_name)
            self.assertEqual(len(expenses), 2)
            self.assertEqual(expenses[0].paid, 0)
            self.assertEqual(expenses[1].paid, 1)

        def test_load_all_from_excel(self):
            clients, sales, workers, expenses = LoadAllFromTable.load_all_from_excel(self.excel_file_name)
            self.assertEqual(len(clients), 2)
            self.assertEqual(clients[0].name, 'Mario')
            self.assertEqual(clients[1].name, 'María')

            self.assertEqual(len(sales), 2)
            self.assertIsNone(sales[0].client_id)
            self.assertIsNone(sales[1].client_id)

            self.assertEqual(len(workers), 2)
            self.assertEqual(workers[0].salary, 6000)
            self.assertEqual(workers[1].salary, 150)

            self.assertEqual(len(expenses), 2)
            self.assertEqual(expenses[0].paid, 0)
            self.assertEqual(expenses[1].paid, 1)

    class TestSaveOneInExcel(unittest.TestCase):
        def setUp(self):
            self.test_excel_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
            self.excel_file_name = self.test_excel_file.name

            wb = Workbook()
            wb.create_sheet(title="Clients")
            wb.create_sheet(title="Sales")
            wb.create_sheet(title="Workers")
            wb.create_sheet(title="Expenses")
            wb.remove(wb["Sheet"])  
            wb.save(self.excel_file_name)

            self.test_client = Client('70000-1', 'María', 'Liendra', 'F', 32, 'liendma@gmail.com')
            self.test_sale = Sale('39000000E', None, '2023-12-10 21:45:00', 15, 0, 1, 'X1 Bowl of Nuggets')
            self.test_worker = Worker('00000-1', 'Viviana', 'Marcic', 'Executive', 6000, 'v.marcic@gmail.com')
            self.test_expense = Expense('45000000W', None, '2022-01-23 21:45:00', 50, 1, 'Buy knifes')

        def tearDown(self):
            self.test_excel_file.close()
            try:
                os.remove(self.excel_file_name)
            except PermissionError:
                pass

        def test_save_client_in_excel(self):
            SaveOneInExcel.save_client_in_excel(self.excel_file_name, self.test_client)
            new_client_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Clients', engine='openpyxl')
            self.assertEqual(len(new_client_in_excel), 1)
            self.assertEqual(new_client_in_excel.loc[0, 'name'], 'María')

        def test_save_sale_in_excel(self):
            SaveOneInExcel.save_sale_in_excel(self.excel_file_name, self.test_sale)
            new_sale_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Sales', engine='openpyxl')
            client_id = new_sale_in_excel['client_id'].apply(lambda x: None if pd.isna(x) else x)
            self.assertEqual(len(new_sale_in_excel), 1)
            self.assertIsNone(client_id[0])

        def test_save_worker_in_excel(self):
            SaveOneInExcel.save_worker_in_excel(self.excel_file_name, self.test_worker)
            new_worker_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Workers', engine='openpyxl')
            self.assertEqual(len(new_worker_in_excel), 1)
            self.assertEqual(new_worker_in_excel.loc[0, 'salary'], 6000)

        def test_save_expense_in_excel(self):
            SaveOneInExcel.save_expense_in_excel(self.excel_file_name, self.test_expense)
            new_expense_in_excel = pd.read_excel(self.excel_file_name, sheet_name='Expenses', engine='openpyxl')
            self.assertEqual(len(new_expense_in_excel), 1)
            self.assertEqual(new_expense_in_excel.loc[0, 'paid'], 1)

class TestInfo:
    class TestInfoClients(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Clients.create_table_client(self.db_file_name)
            Sales.create_table_sale(self.db_file_name)
            Clients.add_client('ABCDE-1', 'John', 'Doe', 'M', 30, 'tr@gmail.com', self.db_file_name)
            Clients.add_client('XYZ00-2', 'Alice', 'Smith', 'F', 25, 'y@gmail.com', self.db_file_name)
            Clients.add_client('12345-3', 'Michael', 'Johnson', 'M', 40, 'urs@gmail.com', self.db_file_name)
            Sales.add_sale('00011123K', 'ABCDE-1', '2024-05-10 21:30:00', 150, 1, 1, 'X5 Pasta Bolognese Combo', self.db_file_name)
            Sales.add_sale('70770707K', '12345-3', '2024-01-01 12:00:00', 50, 0, 1, 'X1 Large Neapolitan Pizza & X1 Large Jam Pizza', self.db_file_name)
            Sales.add_sale('06440400G', '12345-3', '2024-02-04 12:30:00', 10, 0, 0, 'X1 Caprese Salad', self.db_file_name)
            Sales.add_sale('10293847H', 'ABCDE-1', '2024-03-15 18:45:00', 200, 1, 1, 'X2 BBQ Chicken Wings Combo', self.db_file_name)
            Sales.add_sale('99887766L', '12345-3', '2024-03-20 19:15:00', 75, 0, 1, 'X1 Veggie Delight Pizza', self.db_file_name)
            Sales.add_sale('55667788M', 'ABCDE-1', '2024-04-22 13:50:00', 120, 1, 0, 'X3 Spaghetti Carbonara', self.db_file_name)
            Sales.add_sale('22334455N', '12345-3', '2024-04-11 20:10:00', 60, 0, 1, 'X1 Margherita Pizza', self.db_file_name)
            Sales.add_sale('33445566P', 'ABCDE-1', '2024-05-05 14:25:00', 90, 1, 1, 'X2 Caesar Salad & X2 Garlic Bread', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_gender_of_clients(self):
            gender_of_clients = InfoClients.gender_of_clients(self.db_file_name)
            expected = {'Male': 2, 'Female': 1, 'Other': 0}

            self.assertEqual(gender_of_clients, expected)

        def test_age_of_clients(self):
            age_of_clients = InfoClients.age_of_clients(self.db_file_name)
            expected = {'15-20': 0, '21-25': 1, '26-30': 1, '31-35': 0, '36-40': 1, '41-45': 0, '46+': 0}

            self.assertEqual(age_of_clients, expected)

        def test_client_summary(self):
            client_summary_1 = InfoClients.client_summary('ABCDE-1', self.db_file_name)
            client_summary_2 = InfoClients.client_summary('XYZ00-2', self.db_file_name)
            client_summary_3 = InfoClients.client_summary('12345-3', self.db_file_name)

            id_of_history_1 = set([sale.id for sale in client_summary_1['history']])
            expected_history_ids_1 = {'00011123K', '10293847H', '55667788M', '33445566P'}
            id_of_history_2 = set([sale.id for sale in client_summary_2['history']])
            expected_history_ids_2 = set()
            id_of_history_3 = set([sale.id for sale in client_summary_3['history']])
            expected_history_ids_3 = {'70770707K', '06440400G', '99887766L', '22334455N'}

            ids_of_pending_pay_1 = set([sale.id for sale in client_summary_1['pending_pay']])
            expected_pending_pay_ids_1 = set()
            ids_of_pending_pay_2 = set([sale.id for sale in client_summary_2['pending_pay']])
            expected_pending_pay_ids_2 = set()
            ids_of_pending_pay_3 = set([sale.id for sale in client_summary_3['pending_pay']])
            expected_pending_pay_ids_3 = {'70770707K', '06440400G', '99887766L', '22334455N'}

            ids_of_pending_deliver_1 = set([sale.id for sale in client_summary_1['pending_deliver']])
            expected_pending_deliver_ids_1 = {'55667788M'}
            ids_of_pending_deliver_2 = set([sale.id for sale in client_summary_2['pending_deliver']])
            expected_pending_deliver_ids_2 = set()
            ids_of_pending_deliver_3 = set([sale.id for sale in client_summary_3['pending_deliver']])
            expected_pending_deliver_ids_3 = {'06440400G'}

            self.assertEqual(id_of_history_1, expected_history_ids_1)
            self.assertEqual(id_of_history_2, expected_history_ids_2)
            self.assertEqual(id_of_history_3, expected_history_ids_3)
            self.assertEqual(client_summary_1['cash'], 560)
            self.assertEqual(client_summary_2['cash'], 0)
            self.assertEqual(client_summary_3['cash'], 195)
            self.assertEqual(client_summary_1['amount'], 4)
            self.assertEqual(client_summary_2['amount'], 0)
            self.assertEqual(client_summary_3['amount'], 4)
            self.assertEqual(client_summary_1['debt'], 0)
            self.assertEqual(client_summary_2['debt'], 0)
            self.assertEqual(client_summary_3['debt'], 195)
            self.assertEqual(ids_of_pending_pay_1, expected_pending_pay_ids_1)
            self.assertEqual(ids_of_pending_pay_2, expected_pending_pay_ids_2)
            self.assertEqual(ids_of_pending_pay_3, expected_pending_pay_ids_3)
            self.assertEqual(ids_of_pending_deliver_1, expected_pending_deliver_ids_1)
            self.assertEqual(ids_of_pending_deliver_2, expected_pending_deliver_ids_2)
            self.assertEqual(ids_of_pending_deliver_3, expected_pending_deliver_ids_3)

    class TestInfoSales(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()
            Clients.create_table_client(self.db_file_name)
            Sales.create_table_sale(self.db_file_name)
            Sales.add_sale('00011123K', None, '2024-05-10 21:30:00', 150, 1, 1, 'X5 Pasta Bolognese Combo', self.db_file_name)
            Sales.add_sale('70770707K', None, '2024-01-01 12:00:00', 50, 0, 1, 'X1 Large Neapolitan Pizza & X1 Large Jam Pizza', self.db_file_name)
            Sales.add_sale('06440400G', None, '2024-02-04 12:30:00', 10, 0, 0, 'X1 Caprese Salad', self.db_file_name)
            Sales.add_sale('05330401H', None, '2024-03-01 14:15:00', 15, 0, 0, 'X2 Margherita Pizza', self.db_file_name)
            Sales.add_sale('06220402J', None, '2024-03-15 16:45:00', 20, 0, 0, 'X3 Spaghetti Bolognese', self.db_file_name)
            Sales.add_sale('05110403K', None, '2024-04-25 11:00:00', 8, 0, 0, 'X4 Caesar Salad', self.db_file_name)
            Sales.add_sale('06000404L', None, '2024-04-25 13:30:00', 12, 0, 0, 'X5 Pepperoni Pizza', self.db_file_name)
            Sales.add_sale('05990405M', None, '2024-05-05 15:20:00', 18, 0, 0, 'X6 Lasagna', self.db_file_name)
            Sales.add_sale('04820315L', None, '2024-05-10 14:30:00', 22, 1, 0, 'X3 Tiramisu', self.db_file_name)
            Sales.add_sale('03450826H', None, '2024-06-15 18:00:00', 35, 0, 1, 'X4 Chicken Alfredo', self.db_file_name)
            Sales.add_sale('02010607K', None, '2024-06-20 12:45:00', 50, 1, 1, 'X2 Seafood Risotto', self.db_file_name)
            Sales.add_sale('07650904N', None, '2024-05-25 19:10:00', 27, 0, 0, 'X2 Cannoli', self.db_file_name)
            Sales.add_sale('09980712P', None, '2024-07-05 16:40:00', 45, 1, 1, 'X5 Four Cheese Pizza', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_total_summary(self):
            summary = InfoSales.total_summary(self.db_file_name)

            ids_of_pending_transactions = set([sale.id for sale in summary['pending_transactions']])
            ids_of_pending_services = set([sale.id for sale in summary['pending_services']])
            ids_of_all_sales = set([sale.id for sale in summary['all_sales']])

            expected_transaction_states_ids = {'70770707K', '06440400G', '05330401H', '06220402J', '05110403K', '06000404L', '05990405M', '03450826H', '07650904N'}
            expected_transaction_states = {'paid': 4, 'pending': 9}
            expected_service_states = {'delivered': 5, 'pending': 8}
            expected_services_states_ids = {'06440400G', '05330401H', '06220402J', '05110403K', '06000404L', '05990405M', '04820315L', '07650904N'}
            expected_all_ids = {'00011123K', '70770707K', '06440400G', '05330401H', '06220402J', '05110403K', '06000404L', '05990405M', '03450826H', '07650904N', '04820315L', '02010607K', '09980712P'}

            self.assertEqual(summary['cash'], 462)
            self.assertEqual(summary['transaction_states'], expected_transaction_states)
            self.assertEqual(ids_of_pending_transactions, expected_transaction_states_ids)
            self.assertEqual(summary['service_states'], expected_service_states)
            self.assertEqual(ids_of_pending_services, expected_services_states_ids)
            self.assertEqual(ids_of_all_sales, expected_all_ids)
            self.assertEqual(summary['total_sales'], 13)

        def test_daily_summary(self):
            day_1 = datetime(2024, 7, 5)
            summary_1 = InfoSales.dialy_summary(day_1, self.db_file_name)
            day_2 = datetime(2024, 4, 25)
            summary_2 = InfoSales.dialy_summary(day_2, self.db_file_name)
            day_3 = datetime(2023, 2, 1)
            summary_3 = InfoSales.dialy_summary(day_3, self.db_file_name)

            ids_pending_transaction_states_1 = set([sale.id for sale in summary_1['pending_transactions']])
            ids_pending_transaction_states_2 = set([sale.id for sale in summary_2['pending_transactions']])
            ids_pending_transaction_states_3 = set([sale.id for sale in summary_3['pending_transactions']])
            ids_pending_service_states_1 = set([sale.id for sale in summary_1['pending_services']])
            ids_pending_service_states_2 = set([sale.id for sale in summary_2['pending_services']])
            ids_pending_service_states_3 = set([sale.id for sale in summary_3['pending_services']])
            ids_of_all_sales_1 = set([sale.id for sale in summary_1['all_sales']])
            ids_of_all_sales_2 = set([sale.id for sale in summary_2['all_sales']])
            ids_of_all_sales_3 = set([sale.id for sale in summary_3['all_sales']])

            expected_transaction_states_1 = {'paid': 1, 'pending': 0}
            expected_transaction_states_2 = {'paid': 0, 'pending': 2}
            expected_transaction_states_3 = {'paid': 0, 'pending': 0}
            expected_pending_transaction_ids_1 = set()
            expected_pending_transaction_ids_2 = {'05110403K', '06000404L'}
            expected_pending_transaction_ids_3 = set()
            expected_service_states_1 = {'delivered': 1, 'pending': 0}
            expected_service_states_2 = {'delivered': 0, 'pending': 2}
            expected_service_states_3 = {'delivered': 0, 'pending': 0}
            expected_pending_services_ids_1 = set()
            expected_pending_services_ids_2 = {'05110403K', '06000404L'}
            expected_pending_services_ids_3 = set()
            expected_all_sales_id_1 = {'09980712P'}
            expected_all_sales_id_2 = {'05110403K', '06000404L'}
            expected_all_sales_id_3 = set()

            self.assertEqual(summary_1['cash'], 45)
            self.assertEqual(summary_2['cash'], 20)
            self.assertEqual(summary_3['cash'], 0)
            self.assertEqual(summary_1['transaction_states'], expected_transaction_states_1)
            self.assertEqual(summary_2['transaction_states'], expected_transaction_states_2)
            self.assertEqual(summary_3['transaction_states'], expected_transaction_states_3)
            self.assertEqual(ids_pending_transaction_states_1, expected_pending_transaction_ids_1)
            self.assertEqual(ids_pending_transaction_states_2, expected_pending_transaction_ids_2)
            self.assertEqual(ids_pending_transaction_states_3, expected_pending_transaction_ids_3)
            self.assertEqual(summary_1['service_states'], expected_service_states_1)
            self.assertEqual(summary_2['service_states'], expected_service_states_2)
            self.assertEqual(summary_3['service_states'], expected_service_states_3)
            self.assertEqual(ids_pending_service_states_1, expected_pending_services_ids_1)
            self.assertEqual(ids_pending_service_states_2, expected_pending_services_ids_2)
            self.assertEqual(ids_pending_service_states_3, expected_pending_services_ids_3)
            self.assertEqual(ids_of_all_sales_1, expected_all_sales_id_1)
            self.assertEqual(ids_of_all_sales_2, expected_all_sales_id_2)
            self.assertEqual(ids_of_all_sales_3, expected_all_sales_id_3)
            self.assertEqual(summary_1['total_sales'], 1)
            self.assertEqual(summary_2['total_sales'], 2)
            self.assertEqual(summary_3['total_sales'], 0)

class TestAuxiliars(unittest.TestCase):
        def setUp(self):
            self.test_db_file = NamedTemporaryFile(delete=False)
            self.db_file_name = self.test_db_file.name
            self.test_db_file.close()

            Clients.create_table_client(self.db_file_name)
            Clients.add_client('AB7DE-1', 'John', 'Doe', 'M', 30, 'tr@gmail.com', self.db_file_name)

            Sales.create_table_sale(self.db_file_name)
            Sales.add_sale('00000876L', None, '2024-06-20 21:15:00', 80, 1, 1, 'X5 Pasta Bolognese Combo', self.db_file_name)

            Workers.create_table_worker(self.db_file_name)
            Workers.add_worker('OPL67-5', 'Carlos', 'Quintero', 'Cashier', 550, 'plm@gmail.com', self.db_file_name)

            Expenses.create_table_expense(self.db_file_name)
            Expenses.add_expense('98052000B', None, '2008-02-03 20:21:00', 30, 0, 'Recompose a client', self.db_file_name)

        def tearDown(self):
            try:
                os.remove(self.db_file_name)
            except PermissionError:
                pass

        def test_valid_id(self):
            list_clients = Clients.load_clients(self.db_file_name)
            list_sales = Sales.load_sales(self.db_file_name)
            list_workers = Workers.load_workers(self.db_file_name)
            list_expenses = Expenses.load_expenses(self.db_file_name)

            self.assertTrue(valid_id('3456W-W', list_clients, 'client_id_pattern'))
            self.assertFalse(valid_id('75480', list_clients, 'client_id_pattern'))
            self.assertFalse(valid_id('04G5s-a', list_clients, 'client_id_pattern'))
            self.assertFalse(valid_id('AB7DE-1', list_clients, 'client_id_pattern'))

            self.assertTrue(valid_id('00000719J', list_sales, 'sale_id_pattern'))
            self.assertFalse(valid_id('6676U650K', list_sales, 'sale_id_pattern'))
            self.assertFalse(valid_id('3180M', list_sales, 'sale_id_pattern'))
            self.assertFalse(valid_id('00000876L', list_sales, 'sale_id_pattern'))

            self.assertTrue(valid_id('00004-Y', list_workers, 'worker_id_pattern'))
            self.assertFalse(valid_id('870000', list_workers, 'worker_id_pattern'))
            self.assertFalse(valid_id('a0Fj5-0', list_workers, 'worker_id_pattern'))
            self.assertFalse(valid_id('OPL67-5', list_workers, 'worker_id_pattern'))

            self.assertTrue(valid_id('88911000F', list_expenses, 'expense_id_pattern'))
            self.assertFalse(valid_id('992YT890L', list_expenses, 'expense_id_pattern'))
            self.assertFalse(valid_id('8600H', list_expenses, 'expense_id_pattern'))
            self.assertFalse(valid_id('98052000B', list_expenses, 'expense_id_pattern'))